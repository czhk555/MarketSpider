# V2.0-beta1
import csv
import json
import sys
import time
import tkinter
import os
import tkinter.font
import tkinter.ttk as ttk
import traceback
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import font
from threading import Thread
from tkinter import messagebox
from tkinter.ttk import Progressbar
from typing import Any, Dict, Iterable, List, Union
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement


class webdriverError(Exception):
    pass


def show_info_box(title: str, text: str):
    messagebox.showinfo(f"提示 - {title}", text)


def show_warn_box(title: str, text: str):
    messagebox.showwarning(f"警告 - {title}", text)


def show_error_box(title: str, text: str):
    messagebox.showerror(f"错误 - {title}", text)


class EmptyElement:
    text = "?"

    def get_attribute(self, p: str = ""):
        pass


class Gui:
    COLOR_RUNNING = "#3edf86"
    COLOR_DEFAULT = "#eeeeee"
    COLOR_ATTENTION = "#fdc285"
    COLOR_ERROR = "#f23b19"

    def __init__(self, comp_name):
        self.new_window = None
        self.input_box = None
        self.ui_root = tkinter.Tk()
        self.ui_root.title(f"{comp_name} - MarketSpider")
        self.ui_root.geometry("-50+20")
        self.ui_root.attributes("-topmost", 1)
        self.ui_root["background"] = "#eeeeee"
        self.main_text = tkinter.Label(
            self.ui_root,
            text="程序正在准备",
            font=("微软雅黑", 20),
            width=20,
            height=2,
            wraplength=500,
            justify=tkinter.LEFT,
        )
        self.main_text.pack()
        self.main_text["background"] = "#4eb5fb"
        self.status_text = tkinter.Label(
            self.ui_root, text="", relief=tkinter.SUNKEN, anchor="w"
        )
        self.status_text.pack(fill=tkinter.X, side=tkinter.BOTTOM)
        self.progress_bar = Progressbar(self.ui_root)
        self.progress_bar.pack(fill=tkinter.X, side=tkinter.BOTTOM)

    def set_progress(self, done: int, total=100):
        if total == 0:
            self.progress_bar["maximum"] = 50
            self.progress_bar["mode"] = "indeterminate"
            self.progress_bar.start()
        else:
            self.progress_bar["mode"] = "determinate"
            self.progress_bar.stop()
            self.progress_bar["value"] = done
            self.progress_bar["maximum"] = total
            self.ui_root.update()

    def set_status(self, text: str):
        self.status_text["text"] = text
        self.ui_root.update()

    def set_text(self, text: str, color=COLOR_DEFAULT):
        self.main_text["text"] = text
        self.main_text["background"] = color
        self.ui_root.update()

    def save_asked_string(self):
        self.new_window.destroy()  # type: ignore

    def ask_string(self, title: str, description: str):
        """
        弹出新窗口获取字符串
        @param title 新窗口标题
        @param description 新窗口显示的提示文本
        @return 返回用户输入值String
        """
        x = int((self.ui_root.winfo_screenwidth() - self.ui_root.winfo_reqwidth()) / 2)
        y = int(
            (self.ui_root.winfo_screenheight() - self.ui_root.winfo_reqheight()) / 2
        )
        self.new_window = tkinter.Toplevel(master=self.ui_root)
        self.new_window.attributes("-topmost", "true")
        self.new_window.title(title)
        self.new_window.geometry("+{}+{}".format(x, y))
        self.new_window.attributes("-toolwindow", 2)
        tkinter.Label(self.new_window, text=description, anchor="w").pack(
            fill=tkinter.X, expand=True
        )
        v = tkinter.StringVar()
        input_box = tkinter.Entry(self.new_window, textvariable=v)
        input_box.pack()
        tkinter.Button(
            self.new_window, text="确认", command=self.save_asked_string
        ).pack()
        self.ui_root.wait_window(self.new_window)
        return v.get()

    def ui_loop(self):
        self.ui_root.mainloop()

    def ui_start(self):
        Thread(target=self.ui_loop, daemon=True).start()

    def ask_output_formats(self, default_formats: List[str] = None) -> List[str]:
        """
        弹出多选框窗口，让用户选择输出格式
        @param default_formats: 默认选中的格式列表（从 config.json 读取）
        @return: 用户选择的格式列表（["csv", "json", "xlsx", "txt", "sql"] 的子集）
        """
        if default_formats is None:
            default_formats = ["csv", "json"]

        formats = [
            ("CSV", "csv"),
            ("JSON", "json"),
            ("Excel", "xlsx"),
            ("TXT", "txt"),
            ("SQL", "sql")
        ]
        selected_formats = []

        # 创建新窗口
        x = int((self.ui_root.winfo_screenwidth() - self.ui_root.winfo_reqwidth()) / 2)
        y = int((self.ui_root.winfo_screenheight() - self.ui_root.winfo_reqheight()) / 2)
        new_window = tkinter.Toplevel(master=self.ui_root)
        new_window.attributes("-topmost", "true")
        new_window.title("选择输出格式")
        new_window.geometry("+{}+{}".format(x, y))
        new_window.attributes("-toolwindow", 2)

        tkinter.Label(new_window, text="请选择输出文件格式（可多选）：").pack(fill=tkinter.X, padx=5, pady=5)

        # 创建多选框
        vars_dict = {}
        for name, format_id in formats:
            var = tkinter.BooleanVar(value=format_id in default_formats)
            vars_dict[format_id] = var
            tkinter.Checkbutton(new_window, text=name, variable=var).pack(anchor="w", padx=10)

        # 确认按钮
        def confirm():
            nonlocal selected_formats
            selected_formats = [fmt for fmt, var in vars_dict.items() if var.get()]
            if not selected_formats:
                messagebox.showerror("错误", "请至少选择一种输出格式！")
                return
            new_window.destroy()

        tkinter.Button(new_window, text="确认", command=confirm).pack(pady=5)
        self.ui_root.wait_window(new_window)
        return selected_formats


class Logger:
    def __init__(self, comp: str):
        """
        日志记录器
        @param comp 记录根名称
        """
        self.comp = comp
        if not os.path.exists("logs"):
            os.mkdir("Logs")
        log_file_name = time.strftime("%Y_%m_%d", time.localtime())
        self.log_file = open(f"logs\\{log_file_name}.log", "a", encoding="utf8")
        self.write_info("program start...")

    def write_info(self, info: str, dialog=False):
        happened_time = time.strftime("%H:%M:%S", time.localtime())
        self.log_file.write(f"[INFO] {happened_time} [{self.comp}] {info}\n")
        self.log_file.flush()
        if dialog:
            messagebox.showinfo(f"提示 - {self.comp}", info)

    def write_warn(self, warn: str, dialog=False):
        """
        @param warn 友好提示文本
        @param dialog 是否弹出窗体
        """
        err_exc = traceback.format_exc()
        err_stack = traceback.format_stack()
        happened_time = time.strftime("%H:%M:%S", time.localtime())
        self.log_file.write(
            f"[WARN] {happened_time} [{self.comp}] {warn}\n ->err:{err_exc}\n ->stack:"
        )
        for s in err_stack:
            self.log_file.write(s)
            self.log_file.flush()
        self.log_file.flush()
        if dialog:
            messagebox.showwarning(f"注意 - {self.comp}", warn)

    def write_error(self, error_text: str = "", dialog=False):
        """
        记录一个错误.函数自动写入异常和调用链
        @param error_text 友好提示文本
        @param dialog 是否弹出窗体
        """
        err_exc = traceback.format_exc()
        err_stack = traceback.format_stack()
        happened_time = time.strftime("%H:%M:%S", time.localtime())
        self.log_file.write(
            f"[ERROR] {happened_time} 组件:[{self.comp}] 提示文本{error_text}\n{err_exc}"
        )
        self.log_file.flush()
        for s in err_stack:
            self.log_file.write(s)
            self.log_file.flush()
        if dialog:
            messagebox.showerror(f"出错 - {self.comp}", f"{error_text}\n{err_exc}")


class BrowserControl:
    def __init__(self, browser, gui: Gui, ilog: Logger):
        """
        设置浏览器
        @param browser: [chrome,edge,firefox]浏览器类型
        """
        self.gui = gui
        self.ilog = ilog
        gui.set_text("正在创建浏览器实例")
        gui.set_status("正在创建浏览器实例...")
        if browser not in ["chrome", "edge", "firefox"]:
            raise RuntimeError("设置WebDriver失败,目前仅支持Chrome、Edge、Firefox")
        if browser == "chrome":
            browser_option = webdriver.ChromeOptions()
            browser_option.add_argument("--disable-blink-features=AutomationControlled")
            browser_option.add_experimental_option(
                "excludeSwitches", ["enable-logging"]
            )
            self.browser = webdriver.Chrome(options=browser_option)
            self.browser.implicitly_wait(8)

        if browser == "edge":
            browser_option = webdriver.EdgeOptions()
            browser_option.add_argument("--disable-blink-features=AutomationControlled")
            browser_option.add_experimental_option(
                "excludeSwitches", ["enable-logging"]
            )
            self.browser = webdriver.Edge(options=browser_option)
            self.browser.implicitly_wait(8)

        if browser == "firefox":
            browser_option = webdriver.FirefoxOptions()
            browser_option.add_argument("--disable-blink-features=AutomationControlled")
            self.browser = webdriver.Firefox(options=browser_option)
            self.browser.implicitly_wait(8)

    def navi_to(self, url):
        self.gui.set_status('正在前往网页...')
        self.browser.get(url)

    def inject_cookie(self, market):
        self.gui.set_text("正在清空Cookie...")
        self.browser.delete_all_cookies()
        try:
            self.gui.set_text("正在注入Cookie...")
            with open(f"cookie\\{market}.cookie", "r") as f:
                cookie_list = json.load(f)
                for cookie in cookie_list:
                    try:
                        self.browser.add_cookie(cookie)
                    except Exception as e:
                        self.ilog.write_warn(
                            f"inject cookie error index:{cookie}.\n->Err:{e}"
                        )
            self.gui.set_text("正在注入localStorage,sessionStorage...")
            with open(f"cookie\\{market}.storage", "r") as f:
                list = json.load(f)
                ls = json.loads(list["localStorage"])
                ss = json.loads(list["sessionStorage"])
                for key, value in ls.items():
                    try:
                        self.browser.execute_script(
                            f"localStorage.setItem('{key}','{value}')"
                        )
                    except Exception as e:
                        self.ilog.write_warn(
                            f"inject localStorage error index:{key}.\n->Err:{e}"
                        )
                for key, value in ss.items():
                    try:
                        self.browser.execute_script(
                            f"sessionStorage.setItem('{key}','{value}')"
                        )
                    except Exception as e:
                        self.ilog.write_warn(
                            f"inject sessionStorage error index:{key}.\n->Err:{e}"
                        )
        except Exception as e:
            self.ilog.write_warn(
                f"在注入Cookie时出现异常,在抓取时可能因为跳转至登录界面导致爬取失败.\n-> {e}",
                True,
            )
        self.gui.set_text("刷新页面生效Cookie")
        self.browser.refresh()

    def save_cookie(self, market) -> bool:
        try:
            if not os.path.exists("cookie"):
                os.mkdir("cookie")
            with open(f"cookie\\{market}.cookie", "w") as file:
                file.write(json.dumps(self.browser.get_cookies()))
            with open(f"cookie\\{market}.storage", "w") as file:
                storage = {}

                storage["localStorage"] = json.dumps(
                    self.browser.execute_script("return window.localStorage")
                )
                storage["sessionStorage"] = json.dumps(
                    self.browser.execute_script("return window.sessionStorage")
                )
                file.write(json.dumps(storage))
            return False
        except Exception as e:
            print(e)
            return True

    def find_element_css(self, path: str, parent=None) -> Any:
        """查找元素-CSS

        :param path: CSS选择器
        :param parent: (可选)父元素
        :return: 元素引用(异常时返回?)
        """
        try:
            root = self.browser
            if parent != None:
                root = parent
            return root.find_element(By.CSS_SELECTOR, path)
        except:
            print(f"[Browser_XPATH_finder]查找{path}失败")
            self.ilog.write_warn(f"[Browser_CSS_finder]查找{path}失败")
            return EmptyElement()

    def find_element_xpath(self, xpath: str, parent=None) -> Any:
        """查找元素-XPath

        :param xpath: XPath
        :param parent: (可选)父元素
        :return: 元素引用(异常时返回?)
        """
        try:
            root = self.browser
            if parent != None:
                root = parent
            return root.find_element(By.XPATH, xpath)
        except:
            print(f"[Browser_XPATH_finder]查找{xpath}失败")
            self.ilog.write_warn(f"[Browser_CSS_finder]查找{xpath}失败")
            return EmptyElement()

    def scroll_page_end(self) -> None:
        height = 0
        self.gui.set_status('正在翻动页面至底部')
        while height < self.browser.execute_script("return document.body.clientHeight"):
            height += 400
            self.browser.execute_script("window.scrollBy(0, 400)")
            time.sleep(2)

    def setTimeout(self, second: int, reason: str = ""):
        for t in range(second + 1):
            self.gui.set_progress(t, second)
            self.gui.set_status(f"{reason} 正在延时{t}/{second}...")
            time.sleep(1)
    
    def exit(self) -> None:
        self.browser.quit()


class CsvWriter:
    def __init__(self, keyword, market):
        fieldnames = [
            "item_link",
            "item_name",
            "item_price",
            "item_image",
            "item_payment",
            "item_rates",
            "item_sales",
            "item_shop",
            "shop_link",
            "remarks",
        ]
        cn_name = {
            "item_link": "商品链接",
            "item_name": "商品名称",
            "item_price": "商品价格",
            "item_image": "商品图片",
            "item_payment": "已付款人数",
            "item_rates": "已评价数",
            "item_sales": "商品销量",
            "item_shop": "店铺名称",
            "shop_link": "店铺链接",
            "remarks": "备注",
        }
        if not os.path.exists("result"):
            os.mkdir("result")
        self.csvStream = open(
            f'result\\{keyword}-{market}-{time.strftime("%Y-%m-%d_%H-%M", time.localtime())}.csv',
            "a",
            encoding="utf-8-sig",
            newline="",
        )
        self.csvWriter = csv.DictWriter(self.csvStream, fieldnames=fieldnames)
        self.csvWriter.writerow(cn_name)
        self.csvStream.flush()

    def write_new_line(self, tdata: Dict):
        """
        在CSV中插入一条数据
        参数支持以下键:
            item_link:商品链接
            item_name:商品名称
            item_price:商品价格
            item_image:图片链接
            item_payment:商品已支付人数
            item_rates:商品已评价人数
            item_sales:商品销量
            item_shop:店铺名称
            shop_link:店铺链接
            remarks:其他备注信息
        """
        st = {
            "item_link": "",
            "item_name": "",
            "item_price": "",
            "item_image": "",
            "item_payment": "",
            "item_rates": "",
            "item_sales": "",
            "item_shop": "",
            "shop_link": "",
            "remarks": "",
        }
        for key, value in tdata.items():
            st[key] = value
        self.csvWriter.writerow(
            {
                "item_link": st["item_link"],
                "item_name": st["item_name"],
                "item_price": st["item_price"],
                "item_image": st["item_image"],
                "item_payment": st["item_payment"],
                "item_rates": st["item_rates"],
                "item_sales": st["item_sales"],
                "item_shop": st["item_shop"],
                "shop_link": st["shop_link"],
                "remarks": st["remarks"],
            }
        )
        self.csvStream.flush()

    def close_csv(self):
        self.csvStream.close()

class JsonWriter:
    def __init__(self, keyword, market):
        """
        初始化 JSON 文件写入器
        @param keyword: 搜索关键词
        @param market: 平台名称 (如 jd, taobao, 1688)
        """
        if not os.path.exists("result"):
            os.mkdir("result")
        self.filename = f'result/{keyword}-{market}-{time.strftime("%Y-%m-%d_%H-%M", time.localtime())}.json'
        self.data_list = []
        if os.path.exists(self.filename):
            try:
                with open(self.filename, 'r', encoding='utf-8') as f:
                    self.data_list = json.load(f)
            except:
                pass

    def write_new_line(self, tdata: Dict):
        """
        将单条数据追加到 JSON 数据列表中
        @param tdata: 包含商品信息的字典
        """
        self.data_list.append(tdata)
        with open(self.filename, 'w', encoding='utf-8') as f:
            json.dump(self.data_list, f, ensure_ascii=False, indent=2)

    def close_json(self):
        pass


class ExcelWriter:
    def __init__(self, keyword, market):
        """
        初始化 Excel 文件写入器
        @param keyword: 搜索关键词
        @param market: 平台名称 (如 jd, taobao, 1688)
        """
        if not os.path.exists("result"):
            os.mkdir("result")
        self.filename = f'result/{keyword}-{market}-{time.strftime("%Y-%m-%d_%H-%M", time.localtime())}.xlsx'
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "MarketSpider Data"
        # 定义表头
        self.fieldnames = [
            "item_link", "item_name", "item_price", "item_image", "item_payment",
            "item_rates", "item_sales", "item_shop", "shop_link", "remarks"
        ]
        self.cn_names = [
            "商品链接", "商品名称", "商品价格", "商品图片", "已付款人数",
            "已评价数", "商品销量", "店铺名称", "店铺链接", "备注"
        ]
        # 写入表头
        for col, header in enumerate(self.cn_names, 1):
            self.worksheet[f"{get_column_letter(col)}1"] = header
        self.row = 2  # 从第二行开始写入数据

    def write_new_line(self, tdata: Dict):
        """
        将单条数据写入 Excel 文件
        @param tdata: 包含商品信息的字典
        """
        for col, key in enumerate(self.fieldnames, 1):
            self.worksheet[f"{get_column_letter(col)}{self.row}"] = tdata.get(key, "")
        self.row += 1
        self.workbook.save(self.filename)

    def close_excel(self):
        """
        保存并关闭 Excel 文件
        """
        self.workbook.save(self.filename)
        self.workbook.close()


class TxtWriter:
    def __init__(self, keyword, market):
        """
        初始化 TXT 文件写入器
        @param keyword: 搜索关键词
        @param market: 平台名称 (如 jd, taobao, 1688)
        """
        if not os.path.exists("result"):
            os.mkdir("result")
        self.filename = f'result/{keyword}-{market}-{time.strftime("%Y-%m-%d_%H-%M", time.localtime())}.txt'
        self.fieldnames = [
            "item_link", "item_name", "item_price", "item_image", "item_payment",
            "item_rates", "item_sales", "item_shop", "shop_link", "remarks"
        ]
        # 写入表头
        with open(self.filename, 'a', encoding='utf-8') as f:
            f.write("\t".join([
                "商品链接", "商品名称", "商品价格", "商品图片", "已付款人数",
                "已评价数", "商品销量", "店铺名称", "店铺链接", "备注"
            ]) + "\n")

    def write_new_line(self, tdata: Dict):
        """
        将单条数据写入 TXT 文件
        @param tdata: 包含商品信息的字典
        """
        with open(self.filename, 'a', encoding='utf-8') as f:
            values = [str(tdata.get(key, "")) for key in self.fieldnames]
            f.write("\t".join(values) + "\n")

    def close_txt(self):
        pass


class SqlWriter:
    def __init__(self, keyword, market):
        """
        初始化 SQL 文件写入器，生成 INSERT 语句
        @param keyword: 搜索关键词
        @param market: 平台名称 (如 jd, taobao, 1688)
        """
        if not os.path.exists("result"):
            os.mkdir("result")
        self.filename = f'result/{keyword}-{market}-{time.strftime("%Y-%m-%d_%H-%M", time.localtime())}.sql'
        self.table_name = "market_spider_data"
        self.fieldnames = [
            "item_link", "item_name", "item_price", "item_image", "item_payment",
            "item_rates", "item_sales", "item_shop", "shop_link", "remarks"
        ]
        # 写入表结构（可选）
        with open(self.filename, 'a', encoding='utf-8') as f:
            f.write(f"""-- Table structure for {self.table_name}
                CREATE TABLE IF NOT EXISTS {self.table_name} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    item_link TEXT,
                    item_name TEXT,
                    item_price VARCHAR(50),
                    item_image TEXT,
                    item_payment VARCHAR(50),
                    item_rates VARCHAR(50),
                    item_sales VARCHAR(50),
                    item_shop VARCHAR(255),
                    shop_link TEXT,
                    remarks TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;\n\n""")

    def write_new_line(self, tdata: Dict):
        """
        将单条数据写入 SQL 文件作为 INSERT 语句
        @param tdata: 包含商品信息的字典
        """
        values = []
        for key in self.fieldnames:
            value = str(tdata.get(key, "")).replace("'", "''")  # 转义单引号
            values.append(f"'{value}'")
        with open(self.filename, 'a', encoding='utf-8') as f:
            f.write(f"INSERT INTO {self.table_name} ({', '.join(self.fieldnames)}) VALUES ({', '.join(values)});\n")

    def close_sql(self):
        pass


def check_configFile_exist():
    try:
        config_file = open("config.json", "r")
        config_file = json.load(config_file)
        b = config_file["browser"]
        return False
    except:
        return True


def pre_check_configFile():
    try:
        config_file = open("config.json", "r")
        config_file = json.load(config_file)
        b = config_file["browser"]
    except IOError:
        show_error_box(
            "配置文件预检未通过", "未发现配置文件.请运行GetCookie.py获取Cookie."
        )
        sys.exit(1)
    except json.decoder.JSONDecodeError:
        show_error_box(
            "配置文件预检未通过",
            "配置文件不合法.请重新运行GetCookie.py尝试重建配置文件.",
        )
        sys.exit(1)
    except KeyError as err_key:
        show_error_box(
            "配置文件预检未通过",
            f"配置文件缺少关键配置值:{err_key}.\n请重新运行GetCookie.py尝试重建配置文件.",
        )
        sys.exit(1)
    return b


def pre_check_cookie(market):
    try:
        f = open(f"cookie\\{market}.cookie", "r")
        f.close()
        return False
    except FileNotFoundError:
        show_warn_box(
            "未检测到Cookie文件",
            f"未检测到{market}的Cookie文件.\nCookie可以用来存储账户的登录状态，未处于登录状态进行爬取可能出现结果缺失或完全无结果。\n强烈建议您退出本程序,使用GetCookie保存Cookie信息.",
        )
    except PermissionError as pe:
        show_warn_box(
            "检测到Cookie文件异常",
            f"在读取{market}的Cookie文件时出现权限错误{pe}.\n请您检查Cookie文件权限和本程序的运行权限.",
        )
    return True
